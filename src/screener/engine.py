import os
import yaml
import sqlite3
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from src.analytics.cagr import calculate_cagr

def load_screener_config(config_path: str) -> dict:
    """Loads analyst-editable filter thresholds from config/screener_config.yaml."""
    if not os.path.exists(config_path):
        return {}
    with open(config_path, "r") as f:
        config = yaml.safe_load(f)
    return config.get("thresholds", {}) if config else {}

def load_screener_presets(config_path: str) -> dict:
    """Loads preset screener configurations from config/screener_config.yaml."""
    if not os.path.exists(config_path):
        return {}
    with open(config_path, "r") as f:
        config = yaml.safe_load(f)
    return config.get("presets", {}) if config else {}

def calculate_de_score(de: float) -> float:
    """Calculates D/E score with linear interpolation (0 -> 100, 0.5 -> 85, 1 -> 70, 2 -> 50, >5 -> 0)."""
    if de is None or pd.isna(de):
        return 0.0
    if de <= 0:
        return 100.0
    elif de <= 0.5:
        return 100.0 - (de / 0.5) * 15.0
    elif de <= 1.0:
        return 85.0 - ((de - 0.5) / 0.5) * 15.0
    elif de <= 2.0:
        return 70.0 - ((de - 1.0) / 1.0) * 20.0
    elif de <= 5.0:
        return 50.0 - ((de - 2.0) / 3.0) * 50.0
    else:
        return 0.0

def calculate_icr_score(icr: float, label: str = None) -> float:
    """Calculates ICR score with linear interpolation (>10 -> 100, 5 -> 75, 3 -> 50, <1.5 -> 0, Debt Free -> 100)."""
    if label == "Debt Free":
        return 100.0
    if icr is None or pd.isna(icr):
        return 0.0
    if icr < 1.5:
        return 0.0
    elif icr <= 3.0:
        return ((icr - 1.5) / 1.5) * 50.0
    elif icr <= 5.0:
        return 50.0 + ((icr - 3.0) / 2.0) * 25.0
    elif icr <= 10.0:
        return 75.0 + ((icr - 5.0) / 5.0) * 25.0
    else:
        return 100.0

def winsorize_and_scale(series: pd.Series) -> pd.Series:
    """Winsorises a series to P10/P90 and scales it from 0 to 100."""
    vals = series.dropna()
    if len(vals) == 0:
        return pd.Series(0.0, index=series.index)
        
    p10 = np.percentile(vals, 10)
    p90 = np.percentile(vals, 90)
    
    capped = series.clip(lower=p10, upper=p90)
    denom = p90 - p10
    
    if denom == 0:
        return pd.Series(50.0, index=series.index)
    return (capped - p10) / denom * 100

class ScreenerEngine:
    def __init__(self, db_path: str):
        self.db_path = db_path

    def load_base_data(self) -> pd.DataFrame:
        """Loads and merges financial_ratios, market_cap, profitandloss, and sectors into a single DataFrame."""
        if not os.path.exists(self.db_path):
            raise FileNotFoundError(f"Database not found at: {self.db_path}")

        conn = sqlite3.connect(self.db_path)
        
        # Load ratios
        df_ratios = pd.read_sql_query("SELECT * FROM financial_ratios", conn)
        
        # Load market cap
        df_mc = pd.read_sql_query(
            "SELECT company_id, year, market_cap_crore, enterprise_value_crore, "
            "pe_ratio, pb_ratio, ev_ebitda, dividend_yield_pct FROM market_cap", 
            conn
        )
        
        # Load profit and loss (for sales and net profit)
        df_pl = pd.read_sql_query(
            "SELECT company_id, year, sales, net_profit, profit_before_tax, interest FROM profitandloss", 
            conn
        )
        
        # Load sectors and company names
        df_sec = pd.read_sql_query(
            "SELECT c.id as company_id, c.company_name, s.broad_sector, s.sub_sector "
            "FROM companies c "
            "LEFT JOIN sectors s ON c.id = s.company_id", 
            conn
        )
        
        # Load balance sheet (to calculate ROCE on the fly)
        df_bs = pd.read_sql_query(
            "SELECT company_id, year, equity_capital, reserves, borrowings FROM balancesheet", conn
        )
        
        conn.close()
        
        # Merge datasets
        df = pd.merge(df_ratios, df_sec, on="company_id", how="left")
        df = pd.merge(df, df_mc, on=["company_id", "year"], how="left")
        df = pd.merge(df, df_pl, on=["company_id", "year"], how="left")
        df = pd.merge(df, df_bs, on=["company_id", "year"], how="left")
        
        # Load CAGR data to get sales_cagr_3yr (3-year Revenue CAGR)
        from src.analytics.cagr import CAGRCalculationEngine
        project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
        cagr_log = os.path.join(project_root, "output", "ratio_edge_cases.log")
        
        cagr_engine = CAGRCalculationEngine(db_path=self.db_path, log_path=cagr_log)
        df_cagr = cagr_engine.run_cagr_pipeline()
        
        # Select CAGR columns to merge
        df_cagr_sel = df_cagr[["company_id", "year", "sales_cagr_3yr"]].copy()
        df = pd.merge(df, df_cagr_sel, on=["company_id", "year"], how="left")
        
        # Calculate ROCE on the fly
        from src.analytics.sector_roce import calculate_single_roce
        df["computed_roce"] = df.apply(
            lambda r: calculate_single_roce(
                r["profit_before_tax"], r["interest"], r["equity_capital"], r["reserves"], r["borrowings"]
            ), 
            axis=1
        )
        
        # Compute de_declining_yoy
        df = df.sort_values(by=["company_id", "year"])
        df["prev_de"] = df.groupby("company_id")["debt_to_equity"].shift(1)
        df["de_declining_yoy"] = df.apply(
            lambda r: r["debt_to_equity"] < r["prev_de"] if r["debt_to_equity"] is not None and not pd.isna(r["debt_to_equity"]) and r["prev_de"] is not None and not pd.isna(r["prev_de"]) else False,
            axis=1
        )
        
        # Calculate 5-year FCF CAGR
        df_lookup = df.set_index(["company_id", "year"])["free_cash_flow_cr"].to_dict()
        
        def get_lookback_year(year_str: str, lookback: int) -> str | None:
            try:
                parts = year_str.split("-")
                return f"{int(parts[0]) - lookback:04d}-{parts[1]}"
            except Exception:
                return None

        fcf_cagr_list = []
        for _, row in df.iterrows():
            co = row["company_id"]
            yr = row["year"]
            prev_yr = get_lookback_year(yr, 5)
            
            start_fcf = df_lookup.get((co, prev_yr))
            end_fcf = row["free_cash_flow_cr"]
            
            cagr_val, flag = calculate_cagr(start_fcf, end_fcf, 5)
            fcf_cagr_list.append(cagr_val if flag is None else 0.0)
            
        df["fcf_cagr_5yr"] = fcf_cagr_list
        
        # Calculate CFO/PAT ratio
        df["cfo_pat_ratio"] = df.apply(
            lambda r: r["cash_from_operations_cr"] / r["net_profit"] if r["net_profit"] and r["net_profit"] != 0 else 0.0,
            axis=1
        )
        
        # Calculate winsorised Composite Quality Score per broad sector
        df = self.calculate_composite_scores(df)
        
        return df

    def calculate_composite_scores(self, df: pd.DataFrame) -> pd.DataFrame:
        """Calculates sector-relative winsorised Composite Quality Score (Day 17 logic)."""
        df_res = df.copy()
        df_res["composite_quality_score"] = 0.0
        
        # Process each broad sector group
        for sector in df_res["broad_sector"].dropna().unique():
            idx = df_res[df_res["broad_sector"] == sector].index
            df_sector = df_res.loc[idx]
            
            if len(df_sector) == 0:
                continue
                
            # Winsorise and scale relative profitability metrics
            roe_s = winsorize_and_scale(df_sector["return_on_equity_pct"])
            roce_s = winsorize_and_scale(df_sector["computed_roce"])
            npm_s = winsorize_and_scale(df_sector["net_profit_margin_pct"])
            
            # Winsorise and scale relative cash quality metrics
            fcf_cagr_s = winsorize_and_scale(df_sector["fcf_cagr_5yr"])
            cfo_pat_s = winsorize_and_scale(df_sector["cfo_pat_ratio"])
            fcf_pos_s = df_sector["free_cash_flow_cr"].apply(lambda f: 100.0 if f and f > 0 else 0.0)
            
            # Winsorise and scale relative growth metrics
            rev_cagr_s = winsorize_and_scale(df_sector["revenue_cagr_5yr"])
            pat_cagr_s = winsorize_and_scale(df_sector["pat_cagr_5yr"])
            
            # Leverage absolute scores (no winsorisation per specification)
            de_s = df_sector["debt_to_equity"].apply(calculate_de_score)
            icr_s = df_sector.apply(lambda r: calculate_icr_score(r["interest_coverage"], r["icr_label"]), axis=1)
            
            # Dimension weighting
            profitability = 0.15 * roe_s + 0.10 * roce_s + 0.10 * npm_s
            cash_quality = 0.15 * fcf_cagr_s + 0.10 * cfo_pat_s + 0.05 * fcf_pos_s
            growth = 0.10 * rev_cagr_s + 0.10 * pat_cagr_s
            leverage = 0.10 * de_s + 0.05 * icr_s
            
            df_res.loc[idx, "composite_quality_score"] = profitability + cash_quality + growth + leverage
            
        return df_res

    def apply_filters(self, df: pd.DataFrame, thresholds: dict) -> pd.DataFrame:
        """Applies filters to the DataFrame based on specified thresholds."""
        filtered_df = df.copy()
        
        filter_rules = [
            ("roe_min", ">=", "return_on_equity_pct"),
            ("de_max", "<=", "debt_to_equity"),
            ("fcf_min", ">=", "free_cash_flow_cr"),
            ("revenue_cagr_5yr_min", ">=", "revenue_cagr_5yr"),
            ("pat_cagr_5yr_min", ">=", "pat_cagr_5yr"),
            ("opm_min", ">=", "operating_profit_margin_pct"),
            ("pe_max", "<=", "pe_ratio"),
            ("pb_max", "<=", "pb_ratio"),
            ("dividend_yield_min", ">=", "dividend_yield_pct"),
            ("icr_min", ">=", "interest_coverage"),
            ("market_cap_min", ">=", "market_cap_crore"),
            ("net_profit_min", ">=", "net_profit"),
            ("eps_cagr_min", ">=", "eps_cagr_5yr"),
            ("asset_turnover_min", ">=", "asset_turnover"),
            ("sales_min", ">=", "sales"),
            ("dividend_payout_max", "<=", "dividend_payout_ratio_pct"),
            ("revenue_cagr_3yr_min", ">=", "sales_cagr_3yr"),
            ("de_declining_yoy", "==", "de_declining_yoy")
        ]
        
        for key, op, col in filter_rules:
            val = thresholds.get(key)
            if val is None or pd.isna(val):
                continue
            
            if key == "de_declining_yoy":
                if val is True:
                    filtered_df = filtered_df[filtered_df["de_declining_yoy"] == True]
                continue
                
            val = float(val)
            
            if key == "de_max":
                mask = (filtered_df["broad_sector"] == "Financials") | \
                       (filtered_df[col].isna()) | \
                       (filtered_df[col] <= val)
                filtered_df = filtered_df[mask]
            elif key == "icr_min":
                mask = (filtered_df["icr_label"] == "Debt Free") | \
                       (filtered_df[col].isna()) | \
                       (filtered_df[col] >= val)
                filtered_df = filtered_df[mask]
            else:
                if op == ">=":
                    filtered_df = filtered_df[filtered_df[col] >= val]
                elif op == "<=":
                    filtered_df = filtered_df[filtered_df[col] <= val]
                    
        if "composite_quality_score" in filtered_df.columns:
            filtered_df = filtered_df.sort_values(by="composite_quality_score", ascending=False)
            
        return filtered_df

    def run_preset(self, df_base: pd.DataFrame, preset_name: str, config_path: str, year: str = "2024-03") -> pd.DataFrame:
        """Loads thresholds for a preset and runs the filters on the base dataset for a specific year."""
        presets = load_screener_presets(config_path)
        thresholds = presets.get(preset_name, {})
        df_yr = df_base[df_base["year"] == year].copy()
        return self.apply_filters(df_yr, thresholds)

    def export_to_excel(self, df_base: pd.DataFrame, config_path: str, output_path: str, year: str = "2024-03"):
        """Generates output/screener_output.xlsx exporting all 6 presets with cell color-coding."""
        presets = load_screener_presets(config_path)
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Color Fills
        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        border_side = Side(border_style="thin", color="D9D9D9")
        data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        columns_to_export = [
            ("Company ID", "company_id"),
            ("Company Name", "company_name"),
            ("Broad Sector", "broad_sector"),
            ("Sub Sector", "sub_sector"),
            ("NPM %", "net_profit_margin_pct"),
            ("OPM %", "operating_profit_margin_pct"),
            ("ROE %", "return_on_equity_pct"),
            ("D/E", "debt_to_equity"),
            ("ICR", "interest_coverage"),
            ("Asset Turnover", "asset_turnover"),
            ("FCF (Cr)", "free_cash_flow_cr"),
            ("CapEx (Cr)", "capex_cr"),
            ("EPS", "earnings_per_share"),
            ("BVPS", "book_value_per_share"),
            ("Dividend Payout %", "dividend_payout_ratio_pct"),
            ("Total Debt (Cr)", "total_debt_cr"),
            ("CFO (Cr)", "cash_from_operations_cr"),
            ("Rev CAGR 5Yr %", "revenue_cagr_5yr"),
            ("PAT CAGR 5Yr %", "pat_cagr_5yr"),
            ("Composite Score", "composite_quality_score")
        ]
        
        # Mapping from filter criteria key to dataframe column name for formatting
        filter_col_mapping = {
            "roe_min": "return_on_equity_pct",
            "de_max": "debt_to_equity",
            "fcf_min": "free_cash_flow_cr",
            "revenue_cagr_5yr_min": "revenue_cagr_5yr",
            "pat_cagr_5yr_min": "pat_cagr_5yr",
            "opm_min": "operating_profit_margin_pct",
            "pe_max": "pe_ratio",
            "pb_max": "pb_ratio",
            "dividend_yield_min": "dividend_yield_pct",
            "icr_min": "interest_coverage",
            "market_cap_min": "market_cap_crore",
            "net_profit_min": "net_profit",
            "eps_cagr_min": "eps_cagr_5yr",
            "asset_turnover_min": "asset_turnover",
            "sales_min": "sales",
            "dividend_payout_max": "dividend_payout_ratio_pct",
            "revenue_cagr_3yr_min": "sales_cagr_3yr"
        }

        for preset_name, thresholds in presets.items():
            df_preset = self.run_preset(df_base, preset_name, config_path, year=year)
            
            # Create sheet
            ws = wb.create_sheet(title=preset_name[:30]) # Excel sheet name limit is 31 chars
            
            # Write Header
            ws.append([col[0] for col in columns_to_export])
            for col_idx in range(1, len(columns_to_export) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = font_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write Data rows
            for _, row in df_preset.iterrows():
                row_data = []
                for _, col_name in columns_to_export:
                    val = row.get(col_name)
                    # Convert NaN to empty string for clean display
                    if val is None or pd.isna(val):
                        row_data.append("")
                    else:
                        row_data.append(val)
                ws.append(row_data)
                
                # Apply styles and colour-coding
                curr_row = ws.max_row
                for col_idx, (_, col_name) in enumerate(columns_to_export, start=1):
                    cell = ws.cell(row=curr_row, column=col_idx)
                    cell.font = font_data
                    cell.border = data_border
                    
                    # Align strings left, numbers right
                    val = row.get(col_name)
                    if isinstance(val, (int, float)) and not pd.isna(val):
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
                        
                    # Color-code active filter cells
                    for filter_key, limit_val in thresholds.items():
                        mapped_col = filter_col_mapping.get(filter_key)
                        if mapped_col == col_name and limit_val is not None:
                            val_to_check = row.get(col_name)
                            
                            # Handle exceptions
                            if col_name == "debt_to_equity" and row.get("broad_sector") == "Financials":
                                cell.fill = green_fill # Skipped D/E for financials is always green
                            elif col_name == "interest_coverage" and row.get("icr_label") == "Debt Free":
                                cell.fill = green_fill # Debt Free is always green
                            else:
                                if val_to_check is None or pd.isna(val_to_check):
                                    cell.fill = red_fill
                                else:
                                    val_float = float(val_to_check)
                                    limit_float = float(limit_val)
                                    
                                    # Operators mapping
                                    if "min" in filter_key or filter_key == "roe_min" or "cagr" in filter_key:
                                        cell.fill = green_fill if val_float >= limit_float else red_fill
                                    elif "max" in filter_key or filter_key == "de_max":
                                        cell.fill = green_fill if val_float <= limit_float else red_fill
                                        
            # Adjust column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"Exported preset screeners to: {output_path}")

if __name__ == "__main__":
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    db = os.path.join(project_root, "data", "nifty100.db")
    cfg = os.path.join(project_root, "config", "screener_config.yaml")
    out = os.path.join(project_root, "output", "screener_output.xlsx")
    
    engine = ScreenerEngine(db_path=db)
    df_base = engine.load_base_data()
    print(f"Loaded base screener data: {len(df_base)} rows.")
    
    engine.export_to_excel(df_base, cfg, out, year="2024-03")
