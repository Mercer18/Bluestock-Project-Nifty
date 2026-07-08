import os
import sqlite3
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from src.screener.engine import ScreenerEngine

def get_company_peer_group(company_id: str, db_path: str) -> str:
    """Returns the peer group name for a given company, or 'No peer group assigned' if not mapped."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT peer_group_name FROM peer_groups WHERE company_id = ?", (company_id,))
    res = cursor.fetchone()
    conn.close()
    if res:
        return res[0]
    return "No peer group assigned"

def calculate_percent_rank(series: pd.Series, invert: bool = False) -> pd.Series:
    """Calculates percentile rank matching Excel's PERCENTRANK.INC behavior."""
    non_nulls = series.dropna()
    n = len(non_nulls)
    if n == 0:
        return pd.Series(np.nan, index=series.index)
    if n == 1:
        return pd.Series(1.0, index=series.index)
        
    if invert:
        ranks = non_nulls.rank(method="min", ascending=False)
    else:
        ranks = non_nulls.rank(method="min", ascending=True)
        
    min_r = ranks.min()
    max_r = ranks.max()
    
    if max_r == min_r:
        res = pd.Series(1.0, index=non_nulls.index)
    else:
        res = (ranks - min_r) / (max_r - min_r)
        
    return res.reindex(series.index)

class PeerPercentileEngine:
    def __init__(self, db_path: str):
        self.db_path = db_path

    def run_peer_pipeline(self) -> int:
        """Computes percentiles for all peer groups and inserts them into peer_percentiles table."""
        # Load unified ratios data
        screener = ScreenerEngine(self.db_path)
        df_base = screener.load_base_data()
        
        # Load peer groups from database
        conn = sqlite3.connect(self.db_path)
        df_peers = pd.read_sql_query("SELECT * FROM peer_groups", conn)
        conn.close()
        
        # Join peer group mapping
        df = pd.merge(df_base, df_peers, on="company_id", how="inner")
        
        # Metrics to rank (10 metrics)
        metrics_mapping = [
            ("ROE", "return_on_equity_pct", False),
            ("ROCE", "computed_roce", False),
            ("NPM", "net_profit_margin_pct", False),
            ("D/E", "debt_to_equity", True),
            ("FCF", "free_cash_flow_cr", False),
            ("PAT CAGR 5Yr", "pat_cagr_5yr", False),
            ("Revenue CAGR 5Yr", "revenue_cagr_5yr", False),
            ("EPS CAGR 5Yr", "eps_cagr_5yr", False),
            ("Interest Coverage", "interest_coverage", False),
            ("Asset Turnover", "asset_turnover", False)
        ]
        
        insert_records = []
        
        # Group by peer group and year
        for (group_name, year), df_sub in df.groupby(["peer_group_name", "year"]):
            df_sub = df_sub.copy()
            
            # Compute percentiles for each metric
            for metric_label, col_name, invert in metrics_mapping:
                if col_name not in df_sub.columns:
                    continue
                    
                pct_ranks = calculate_percent_rank(df_sub[col_name], invert=invert)
                
                for idx, row in df_sub.iterrows():
                    co = row["company_id"]
                    val = row[col_name]
                    rank_val = pct_ranks.loc[idx]
                    
                    db_val = float(val) if val is not None and not pd.isna(val) else None
                    db_rank = float(rank_val) if rank_val is not None and not pd.isna(rank_val) else None
                    
                    insert_records.append((
                        co,
                        group_name,
                        metric_label,
                        db_val,
                        db_rank,
                        year
                    ))
                    
        # Write to SQLite
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM peer_percentiles")
        cursor.executemany(
            "INSERT INTO peer_percentiles (company_id, peer_group_name, metric, value, percentile_rank, year) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            insert_records
        )
        conn.commit()
        conn.close()
        
        print(f"Populated peer_percentiles table with {len(insert_records)} records.")
        return len(insert_records)

    def generate_peer_comparison_report(self, output_path: str, year: str = "2024-03"):
        """Generates output/peer_comparison.xlsx with 11 sheets covering all peer groups."""
        screener = ScreenerEngine(self.db_path)
        df_base = screener.load_base_data()
        
        # Get peer group mapping
        conn = sqlite3.connect(self.db_path)
        df_peers = pd.read_sql_query("SELECT * FROM peer_groups", conn)
        
        # Load peer percentiles for target year
        df_pcts = pd.read_sql_query(
            "SELECT company_id, metric, percentile_rank FROM peer_percentiles WHERE year = ?", 
            conn, params=(year,)
        )
        conn.close()
        
        # Merge peer info
        df_latest = df_base[df_base["year"] == year].copy()
        df_latest = pd.merge(df_latest, df_peers, on="company_id", how="inner")
        
        # Pivot percentile ranks to columns
        df_pct_pivot = df_pcts.pivot(index="company_id", columns="metric", values="percentile_rank")
        df_pct_pivot = df_pct_pivot.add_suffix(" Percentile").reset_index()
        
        # Merge percentiles
        df_merged = pd.merge(df_latest, df_pct_pivot, on="company_id", how="left")
        
        # Create workbook
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Colors
        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        gold_fill = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        median_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        font_bold = Font(name="Calibri", size=11, bold=True)
        
        border_side = Side(border_style="thin", color="D9D9D9")
        data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # 20 metric columns to export
        metrics_cols = [
            ("Sales (Cr)", "sales"),
            ("Net Profit (Cr)", "net_profit"),
            ("NPM %", "net_profit_margin_pct"),
            ("OPM %", "operating_profit_margin_pct"),
            ("ROE %", "return_on_equity_pct"),
            ("ROCE %", "computed_roce"),
            ("D/E", "debt_to_equity"),
            ("ICR", "interest_coverage"),
            ("Asset Turnover", "asset_turnover"),
            ("FCF (Cr)", "free_cash_flow_cr"),
            ("CapEx (Cr)", "capex_cr"),
            ("EPS", "earnings_per_share"),
            ("BVPS", "book_value_per_share"),
            ("Dividend Payout %", "dividend_payout_ratio_pct"),
            ("Total Debt (Cr)", "total_debt_cr"),
            ("CFO (Cr)", "cash_from_operations_cr"),
            ("Rev CAGR 5Yr %", "revenue_cagr_5yr"),
            ("PAT CAGR 5Yr %", "pat_cagr_5yr"),
            ("EPS CAGR 5Yr %", "eps_cagr_5yr"),
            ("P/E", "pe_ratio")
        ]
        
        percentile_cols = [
            ("ROE Percentile", "ROE Percentile"),
            ("ROCE Percentile", "ROCE Percentile"),
            ("NPM Percentile", "NPM Percentile"),
            ("D/E Percentile", "D/E Percentile"),
            ("FCF Percentile", "FCF Percentile"),
            ("PAT CAGR Percentile", "PAT CAGR 5Yr Percentile"),
            ("Rev CAGR Percentile", "Revenue CAGR 5Yr Percentile"),
            ("EPS CAGR Percentile", "EPS CAGR 5Yr Percentile"),
            ("ICR Percentile", "Interest Coverage Percentile"),
            ("Asset Turnover Percentile", "Asset Turnover Percentile")
        ]
        
        columns_to_export = [
            ("Company ID", "company_id"),
            ("Company Name", "company_name")
        ] + metrics_cols + percentile_cols
        
        # Create sheet per peer group
        for group_name, df_group in df_merged.groupby("peer_group_name"):
            df_group = df_group.copy().sort_values(by="company_id")
            
            ws = wb.create_sheet(title=group_name[:30])
            
            # Write Header
            ws.append([col[0] for col in columns_to_export])
            for col_idx in range(1, len(columns_to_export) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = font_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Write Data
            for _, row in df_group.iterrows():
                row_data = []
                for _, col_name in columns_to_export:
                    val = row.get(col_name)
                    row_data.append(val if val is not None and not pd.isna(val) else "")
                ws.append(row_data)
                
                curr_row = ws.max_row
                is_bench = row.get("is_benchmark") == 1 or row.get("is_benchmark") is True
                
                # Apply formatting
                for col_idx, (_, col_name) in enumerate(columns_to_export, start=1):
                    cell = ws.cell(row=curr_row, column=col_idx)
                    cell.font = font_data
                    cell.border = data_border
                    
                    val = row.get(col_name)
                    
                    # Alignments
                    if isinstance(val, (int, float)) and not pd.isna(val):
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
                        
                    # Color-code percentile cells
                    if "Percentile" in col_name and val is not None and not pd.isna(val):
                        val_float = float(val)
                        if val_float >= 0.75:
                            cell.fill = green_fill
                        elif val_float >= 0.25:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill
                    elif is_bench:
                        # Highlight benchmark company row in gold/amber
                        cell.fill = gold_fill
                        
            # Add Summary Median Row
            median_row = ["Median", ""]
            for _, col_name in columns_to_export[2:]:
                # Compute median of numeric values in this column for this group
                vals = df_group[col_name].dropna()
                # Exclude strings (like empty strings)
                numeric_vals = pd.to_numeric(vals, errors="coerce").dropna()
                if len(numeric_vals) > 0:
                    median_row.append(float(np.median(numeric_vals)))
                else:
                    median_row.append("")
                    
            ws.append(median_row)
            curr_row = ws.max_row
            
            # Format Median Row
            for col_idx in range(1, len(columns_to_export) + 1):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = font_bold
                cell.fill = median_fill
                cell.border = data_border
                val = median_row[col_idx - 1]
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
                    
            # Column width auto-sizing
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"Exported peer comparison report to: {output_path}")

if __name__ == "__main__":
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    db = os.path.join(project_root, "data", "nifty100.db")
    out = os.path.join(project_root, "output", "peer_comparison.xlsx")
    
    print("Testing company mapping:")
    print("  INFY ->", get_company_peer_group("INFY", db))
    print("  XYZ (Invalid) ->", get_company_peer_group("XYZ", db))
    
    engine = PeerPercentileEngine(db_path=db)
    engine.run_peer_pipeline()
    engine.generate_peer_comparison_report(output_path=out, year="2024-03")
